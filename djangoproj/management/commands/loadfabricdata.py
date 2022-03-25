import logging
import traceback


from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.db import transaction
from djangoproj.models import SteelStructureFabric, TransportDelivery, DeliveryObject
from djangoproj.models.fabric import TransportStatus

known_attributes_dict = {
    'Дата': 'date',
    'Вес(тн)': 'weight',
    'Статус': 'status',
    'Дата выгрузки': 'unloading_date',
    '№': 'id',
    '№ УПД': 'doc_id',
}

reversed_statuses = {
    choice[1]: choice[0]
    for choice in TransportStatus.choices
}


class Command(BaseCommand):
    help = 'Loads anime data from csv file.'
    start_fabric_margin = [3, 2]
    data_margin = [2, 0]
    ws = None

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs=1, type=str)

    def get_cell(self, row, col, data=False):
        data_margin = self.data_margin if data else [0, 0]
        return self.ws.cell(
            row=row + self.start_fabric_margin[0] + data_margin[0],
            column=col + self.start_fabric_margin[1] + data_margin[1],
        )

    def handle(self, *args, **options):
        wb = load_workbook(filename=options['filename'][0], read_only=True, data_only=True)
        self.ws = wb.worksheets[0]
        self.start_fabric_margin = [3, 2]
        while True:
            fabric_name = self.get_cell(0, 0).value
            if fabric_name is None:
                break
            fabric, _ = SteelStructureFabric.objects.get_or_create(name=fabric_name)
            fabric_rows_count = self.handle_fabric(fabric)
            self.start_fabric_margin[1] += fabric_rows_count + 1

    def handle_fabric(self, fabric):
        headers = self.get_headers()
        row_margin = 0
        while True:
            row_data = self.get_row_data(row_margin, headers)
            if row_data is None:
                break
            row_data_dict = dict(zip(headers, row_data))
            self.handle_delivery(row_data_dict, fabric)
            row_margin += 1

        return len(headers)

    @transaction.atomic
    def handle_delivery(self, data_dict, fabric):
        data_to_delivery = {}
        for data_row in data_dict.items():
            if data_row[0] not in known_attributes_dict:
                continue
            db_field = known_attributes_dict[data_row[0]]
            data_to_delivery[db_field] = data_row[1]

        data_to_delivery.pop('id')
        if 'status' in data_to_delivery:
            status_to_db = reversed_statuses.get(data_to_delivery['status'].strip(), None)
            data_to_delivery['status'] = status_to_db
        try:
            delivery, _ = TransportDelivery.objects.update_or_create(**data_to_delivery, fabric=fabric)
            self.handle_objects(data_dict, delivery)
        except Exception as e:
            logging.error(traceback.format_exc())
            print('An exception during delivery loading occurred: {}'.format(e))

    @staticmethod
    def handle_objects(data_dict, delivery):
        for data_row in data_dict.items():
            if data_row[0] in known_attributes_dict:
                continue
            DeliveryObject.objects.update_or_create(delivery=delivery, name=data_row[0], value=data_row[1])

    def get_headers(self):
        headers = []
        while True:
            data = self.get_cell(0, len(headers), data=True).value
            if data is None:
                break
            headers.append(data)

        return headers

    def get_row_data(self, margin, headers):
        row_data = []
        for i in range(len(headers)):
            row_data.append(self.get_cell(margin + 1, i, data=True).value)
        if row_data[0] is None:
            return None
        return row_data
